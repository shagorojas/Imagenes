
# Importamos las librerías necesarias
from xlsxwriter.utility import xl_rowcol_to_cell, xl_cell_to_rowcol
from collections import defaultdict
from openpyxl.drawing.image import Image as ExcelImage  # Renombramos Image para evitar conflicto con PIL
from PIL import Image as PILImage  # Renombramos también Image de PIL
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from copy import copy  # ✅ Importar copy para clonar estilos
import concurrent.futures
import win32com.client
import pandas as pd
import pythoncom
import numpy as np
import pathlib
import shutil
import time
import os
import re

# Definir rutas
ruta_master = os.path.join(str(os.path.abspath(pathlib.Path().absolute())))
ruta_json = os.path.join(ruta_master, "Config", "Config.json")
ruta_log = os.path.join(ruta_master, "Log", "Eventos.log")
ruta_resultado = os.path.join(ruta_master, "Resultado")
ruta_resultado_pdf = os.path.join(ruta_master, "Resultado pdf")
ruta_resultado_combinado = os.path.join(ruta_master, "Resultado excel")

# Rutas imagenes
logo_alimentos = os.path.join(ruta_master, "util", "Logo alimentos.png")
logo_operador = os.path.join(ruta_master, "util", "Logo operador.png")
logo_secretaria = os.path.join(ruta_master, "util", "Logo secretaria.png")

# Ruta insumos
ruta_archivo_focalizacion = "Insumo\\Focalizacion.xlsx"
ruta_archivo_aplicacion_novedades = "Insumo\\Focalizacion_actualizada.xlsx"
ruta_archivo_novedades = "Insumo\\Novedades.xlsx"
ruta_parametros = "Insumo\\Parametros.xlsx"

# Cargamos los parametros
df_parametros = pd.read_excel(ruta_parametros)

# Convertir a diccionario
dict_data = dict(zip(df_parametros["Concepto"], df_parametros["Valor"]))


# Cargamos los parametros por variables
departamento = dict_data["Departamento"]
municipio = dict_data["Municipio"]
operador = dict_data["Operador"]
contrato = dict_data["Contrato No."]
codigo_dane = dict_data["Codigo dane"]
codigo_dane_completo = dict_data["Codigo dane completo"]
mes_atencion = dict_data["Mes de atencion"]
anio = dict_data["Año"]

class GeneradorPlantillas:
    def __init__(self):
        pass

    def crear_plantilla_control(self, df_plantilla, nombre_df, var_sede, var_codigo_dane_sede, var_jornada, df_encabezado, var_grado):

        # Crear un DataFrame vacío
        df = pd.DataFrame(index=range(10), columns=[chr(65 + i) for i in range(35)])

        # Guardar el DataFrame en un archivo Excel
        archivo_excel = f"Resultado\\{nombre_df}.xlsx"
        df.to_excel(archivo_excel, index=False, engine='xlsxwriter')

        # Crear una conexión con el archivo Excel y agregar las imágenes
        writer = pd.ExcelWriter(archivo_excel, engine='xlsxwriter')
        df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')

        # Acceder al objeto workbook y worksheet
        workbook  = writer.book
        worksheet = writer.sheets['Sheet1']

        # Insertar las imágenes
        worksheet.insert_image('AD1', logo_alimentos, {'x_scale': 0.8, 'y_scale': 0.8})
        worksheet.insert_image('D1', logo_operador, {'x_scale': 0.8, 'y_scale': 0.8})
        worksheet.insert_image('A1', logo_secretaria, {'x_scale': 0.8, 'y_scale': 0.8})

        # Combinar celdas de A8 a AI8
        worksheet.merge_range('A8:AI8', 'Formato - REGISTRO Y CONTROL DIARIO DE ASISTENCIA', 
                                workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'fg_color': 'black',   # Color de fondo negro
                                'font_color': 'white',  # Color de texto blanco
                                'font_size': 12        # Tamaño de fuente
                            }))

        # Combinar celdas A9 y B9 y agregar el texto "DEPARTAMENTO" en negrita
        worksheet.merge_range('A9:B9', 'DEPARTAMENTO:', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_size': 12
                            }))

        # Definir el formato con borde inferior negro
        borde_inferior_negro = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'font_size': 12,
            'bottom': 1,  # Borde inferior negro
        })

        # Definir el formato con borde inferior negro
        borde_inferior = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'font_size': 12,
            'bottom': 1,  # Borde inferior negro
        })

        # Combinar celdas C9 y D9 y agregar el texto de la variable "departamento"
        worksheet.merge_range('C9:D9', departamento, borde_inferior_negro)

        # Combinar celdas A10 y B10 y agregar el texto "MUNICIPIO" en negrita
        worksheet.merge_range('A10:B10', 'MUNICIPIO:', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_size': 12
                            }))

        # Combinar celdas C9 y D9 y agregar el texto de la variable "municipio"
        worksheet.merge_range('C10:D10', municipio , borde_inferior_negro)

        # Combinar celdas A11 y B11 y agregar el texto "OPERADOR" en negrita
        worksheet.merge_range('A11:B11', 'OPERADOR:', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_size': 12
                            }))

        # Combinar celdas C11 y D11 y agregar el texto de la variable "operador"
        worksheet.merge_range('C11:D11', operador , borde_inferior_negro)

        # Combinar celdas A12 y B12 y agregar el texto "CONTRATO No" en negrita
        worksheet.merge_range('A12:B12', 'CONTRATO No:', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_size': 12
                            }))

        # Combinar celdas C12 y D12 y agregar el texto de la variable "contrato"
        worksheet.merge_range('C12:D12', contrato , borde_inferior_negro)

        # Combinar celdas G9 y L9 y agregar el texto "MUNICIPIO" en negrita
        worksheet.merge_range('G9:L9', 'NOMBRE DE INSTITUCIÓN O CENTRO EDUCATIVO:', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_size': 12
                            }))

        # Combinar celdas G10 y L10 y agregar el texto "MUNICIPIO" en negrita
        worksheet.merge_range('G10:L10', 'CODIGO DANE INSTITUCIÓN O CENTRO EDUCATIVO:', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_size': 12
                            }))

        # Combinar celdas G11 y J11 y agregar el texto "MUNICIPIO" en negrita
        worksheet.merge_range('G11:J11', 'MES DE ATENCIÓN:', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_size': 12
                            }))

        # Combinar celdas R11 y S11 y agregar el texto "MUNICIPIO" en negrita
        worksheet.merge_range('R11:S11', 'AÑO:', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_size': 12
                            }))

        # Combinar celdas P9 y AF9 y agregar el texto de la variable "institucion"
        worksheet.merge_range('P9:AF9', var_sede , borde_inferior)
        # Combinar celdas P10 y AF10 y agregar el texto de la variable "institucion"
        worksheet.merge_range('P10:AF10', str(var_codigo_dane_sede) , borde_inferior)
        # Combinar celdas K11 y P11 y agregar el texto de la variable "institucion"
        worksheet.merge_range('K11:P11', mes_atencion , borde_inferior)
        # Combinar celdas T11 y X11 y agregar el texto de la variable "institucion"
        worksheet.merge_range('T11:X11', anio , borde_inferior)


        # Escribir "CÓDIGO DANE:" en E9 y B12 en negrita, sin combinar celdas
        worksheet.write('E9', 'CÓDIGO DANE:', workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'font_size': 12
        }))

        worksheet.write('E10', 'CÓDIGO DANE:', workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'font_size': 12
        }))

        worksheet.write('E11', 'JORNADA:', workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'font_size': 12
        }))

        # Escribir el valor de la variables en las celdas respectivas
        worksheet.write('F9', codigo_dane, borde_inferior)
        worksheet.write('F10', codigo_dane_completo, borde_inferior)
        worksheet.write('F11', var_jornada, borde_inferior)

        # Combinar las celdas A14, A15 y A16, agregar el texto "N°" con el color de fondo #A6A6A6
        worksheet.merge_range('A14:A16', 'N°', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',   # Establecer la fuente como Arial
                                'font_size': 14,        # Tamaño de fuente 14
                                'bg_color': '#A6A6A6'  # Color de fondo
                            }))
        # Combinar las celdas agregar el texto "TIPO DE DOCUMENTO" con el color de fondo #A6A6A6
        worksheet.merge_range('B14:B16', 'TIPO DE DOCUMENTO', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',   # Establecer la fuente como Arial
                                'font_size': 14,        # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo
                                'text_wrap': True       # Ajustar el texto para que se ajuste dentro de la celda
                            }))
        # Combinar las celdas agregar el texto "NÚMERO DE DOCUMENTO DE IDENTIDAD" con el color de fondo #A6A6A6
        worksheet.merge_range('C14:C16', 'NÚMERO DE DOCUMENTO DE IDENTIDAD', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',   # Establecer la fuente como Arial
                                'font_size': 14,        # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo
                                'text_wrap': True       # Ajustar el texto para que se ajuste dentro de la celda
                            }))
        # Combinar las celdas agregar el texto "PRIMER NOMBRE DEL TITULAR DE DERECHO" con el color de fondo #A6A6A6
        worksheet.merge_range('D14:D16', 'PRIMER NOMBRE DEL TITULAR DE DERECHO', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',   # Establecer la fuente como Arial
                                'font_size': 14,        # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo
                                'text_wrap': True       # Ajustar el texto para que se ajuste dentro de la celda
                            }))
        # Combinar las celdas agregar el texto "SEGUNDO NOMBRE DEL TITULAR DE DERECHO" con el color de fondo #A6A6A6
        worksheet.merge_range('E14:E16', 'SEGUNDO NOMBRE DEL TITULAR DE DERECHO', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',   # Establecer la fuente como Arial
                                'font_size': 14,        # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo
                                'text_wrap': True       # Ajustar el texto para que se ajuste dentro de la celda
                            }))
        # Combinar las celdas agregar el texto "PRIMER APELLIDO DEL TITULAR DE DERECHO" con el color de fondo #A6A6A6
        worksheet.merge_range('F14:F16', 'PRIMER APELLIDO DEL TITULAR DE DERECHO', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',   # Establecer la fuente como Arial
                                'font_size': 14,        # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo
                                'text_wrap': True       # Ajustar el texto para que se ajuste dentro de la celda
                            }))
        # Combinar las celdas agregar el texto "SEGUNDO APELLIDO DEL TITULAR DE DERECHO" con el color de fondo #A6A6A6
        worksheet.merge_range('G14:G16', 'SEGUNDO APELLIDO DEL TITULAR DE DERECHO', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',   # Establecer la fuente como Arial
                                'font_size': 14,        # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo
                                'text_wrap': True      # Ajustar texto dentro de la celda
                            }))
        # Combinar las celdas agregar el texto "Sexo" con el color de fondo #A6A6A6
        worksheet.merge_range('H14:H16', 'Sexo', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',   # Establecer la fuente como Arial
                                'font_size': 14,        # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo
                                'rotation': 90          # Rotar el texto 90 grados hacia arriba
                            }))
        # Combinar las celdas agregar el texto "Grado Educativo" con el color de fondo #A6A6A6
        worksheet.merge_range('I14:I16', 'Grado Educativo', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',   # Establecer la fuente como Arial
                                'font_size': 14,        # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo
                                'rotation': 90          # Rotar el texto 90 grados hacia arriba
                            }))

        # Combinar las celdas agregar el texto "Tipo de complemento" con el color de fondo #A6A6A6
        worksheet.merge_range('J14:J16', 'Tipo de complemento', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',   # Establecer la fuente como Arial
                                'font_size': 14,        # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo
                                'rotation': 90,          # Rotar el texto 90 grados hacia arriba
                                'text_wrap': True      # Ajustar texto dentro de la celda
                            }))


        # 1. Desactivar la cuadrícula
        worksheet.hide_gridlines(2)  # 2 es para ocultar la cuadrícula en la vista de diseño

        # 2. Rellenar toda la hoja con color blanco
        formato_blanco = workbook.add_format({'bg_color': 'white'})
        worksheet.set_column('A:Z', None, formato_blanco)  # Rellenar celdas de la A a la Z con color blanco (ajustar según el número de columnas)


        # 3. Ajustar el tamaño de las columnas de acuerdo con un ancho específico
        column_widths = {
            'A': 19,     # Ancho de la columna A
            'B': 19.91,  # Ancho de la columna B
            'C': 35,     # Ancho de la columna C
            'D': 45,     # Ancho de la columna D
            'E': 51.18,  # Ancho de la columna E
            'F': 45.55,  # Ancho de la columna F
            'G': 45.18,  # Ancho de la columna G
            'H': 6.18,   # Ancho de la columna H
            'I': 5.55,   # Ancho de la columna I
            'J': 13.55,  # Ancho de la columna J
            'K': 4.91,   # Ancho de la columna K
            'L': 4.91,   # Ancho de la columna L
            'M': 4.91,   # Ancho de la columna M
            'N': 4.91,   # Ancho de la columna N
            'O': 4.91,   # Ancho de la columna O
            'P': 4.91,   # Ancho de la columna P
            'Q': 4.91,   # Ancho de la columna Q
            'R': 4.91,   # Ancho de la columna R
            'S': 4.91,   # Ancho de la columna S
            'T': 4.91,   # Ancho de la columna T
            'U': 4.91,   # Ancho de la columna U
            'V': 4.91,   # Ancho de la columna V
            'W': 4.91,   # Ancho de la columna W
            'X': 4.91,   # Ancho de la columna X
            'Y': 4.91,   # Ancho de la columna Y
            'Z': 4.91,   # Ancho de la columna Z
            'AA': 4.91,  # Ancho de la columna AA
            'AB': 4.91,  # Ancho de la columna AB
            'AC': 4.91,  # Ancho de la columna AC
            'AD': 4.91,  # Ancho de la columna AD
            'AE': 4.91,  # Ancho de la columna AE
            'AF': 4.91,  # Ancho de la columna AF
            'AG': 4.91   # Ancho de la columna AG
        }

        # Asignar el ancho especificado a cada columna
        for col, width in column_widths.items():
            worksheet.set_column(f'{col}:{col}', width)  # Establecer el ancho para cada columna


        # Combinar las celdas K14:AI14 y agregar el texto con el color de fondo #A6A6A6
        worksheet.merge_range('K14:AI14', 'FECHA DE ENTREGA - Escriba el día hábil al cual corresponde la entrega del Complemento Alimentario', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',  # Establecer la fuente como Arial
                                'font_size': 14,       # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo gris
                                'text_wrap': True      # Ajustar texto dentro de la celda
                            }))

        # Combinar las celdas K16:AG16 y agregar el texto con el color de fondo #A6A6A6
        worksheet.merge_range('K16:AG16', 'Número de días de atención - Marque con una X el día que el Titular de Derecho recibe  el complemento alimentario', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',  # Establecer la fuente como Arial
                                'font_size': 14,       # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo gris
                                'text_wrap': True      # Ajustar texto dentro de la celda
                            }))

        # Combinar las celdas AH15:AI16 y agregar el texto con el color de fondo #A6A6A6
        worksheet.merge_range('AH15:AI16', 'Total días de consumo', 
                            workbook.add_format({
                                'bold': True,          # Negrita
                                'align': 'center',     # Alinear al centro
                                'valign': 'vcenter',   # Alinear verticalmente al centro
                                'font_name': 'Arial',  # Establecer la fuente como Arial
                                'font_size': 14,       # Tamaño de fuente 14
                                'bg_color': '#A6A6A6',  # Color de fondo gris
                                'text_wrap': True      # Ajustar texto dentro de la celda
                            }))

        # Ajustar la altura de las filas
        for fila in range(1, 8):  # Filas de 1 a 7
            worksheet.set_row(fila - 1, 12)  

        worksheet.set_row(7, 18)  # Fila 8 (índice 7 en Python)

        for fila in range(9, 14):  # Filas de 9 a 13
            worksheet.set_row(fila - 1, 30)

        worksheet.set_row(13, 42.8)  # Fila 14 (índice 13 en Python)
        worksheet.set_row(14, 68.3)  # Fila 15 (índice 14 en Python)

        for fila in range(16, 37):  # Filas de 16 a 36
            worksheet.set_row(fila - 1, 36.8)
        
        worksheet.set_row(36, 30)  # Fila 37 (índice 36 en Python)
        worksheet.set_row(37, 30)  # Fila 38 (índice 37 en Python)
        worksheet.set_row(38, 54)  # Fila 39 (índice 38 en Python)
        worksheet.set_row(39, 30)  # Fila 40 (índice 39 en Python)
        worksheet.set_row(40, 30)  # Fila 41 (índice 40 en Python)
        worksheet.set_row(41, 30)  # Fila 42 (índice 41 en Python)
        worksheet.set_row(42, 30)  # Fila 43 (índice 42 en Python)
        worksheet.set_row(43, 30)  # Fila 44 (índice 43 en Python)
        worksheet.set_row(44, 30)  # Fila 45 (índice 41 en Python)
        worksheet.set_row(45, 54)  # Fila 46 (índice 41 en Python)
        worksheet.set_row(46, 30)  # Fila 47 (índice 41 en Python)
        worksheet.set_row(47, 19.5)  # Fila 48 (índice 41 en Python)
        worksheet.set_row(48, 145)  # Fila 49 (índice 41 en Python)
        worksheet.set_row(49, 15)  # Fila 50 (índice 41 en Python)
        worksheet.set_row(50, 30)  # Fila 51 (índice 41 en Python)
        worksheet.set_row(51, 30)  # Fila 52 (índice 41 en Python)

        # Combinar celdas de dos en dos en la columna AH y AI, desde la fila 17 hasta la 36
        merge_format = workbook.add_format({
            'align': 'center',     # Alinear al centro
            'valign': 'vcenter',   # Alinear verticalmente al centro
            'text_wrap': True,     # Ajustar texto
        })

        for row in range(17, 37):  # Desde la fila 17 hasta la 36
            worksheet.merge_range(f'AH{row}:AI{row}', '', merge_format)

        # Lista de celdas combinadas para evitar sobrescribir su formato
        celdas_combinadas = [
            'A14:A16',  # Ejemplo de combinación
            'B14:B16',
            'C14:C16',
            'D14:D16',
            'E14:E16',
            'F14:F16',
            'G14:G16',
            'H14:H16',
            'I14:I16',
            'J14:J16',
            'K14:AI14',
            'K16:AG16',
            'AH15:AI16'
        ]

        # Definir formato con bordes
        border_format_combined = workbook.add_format({
            'border': 1,       # Borde en todas las direcciones
            'bold': True,      # Negrita (opcional, si ya lo usaste en otras celdas)
            'align': 'center',  # Alinear al centro
            'valign': 'vcenter' # Alinear verticalmente al centro
        })


        # Aplicar SOLO el formato con bordes a celdas ya combinadas
        for merge_range in celdas_combinadas:
            worksheet.conditional_format(merge_range, {'type': 'no_errors', 'format': border_format_combined})


        # Formato de borde
        border_format = workbook.add_format({'border': 1})

        # Recorremos las filas y columnas dentro del rango A14:AI36
        for row in range(14, 36):  # De fila 14 a 36
            for col in range(0, 35):  # De columna A (0) hasta AI (34)
                cell_ref = xl_rowcol_to_cell(row, col)  # Convertimos a referencia A1 (Ej: "B15")

                # Si la celda no está dentro de una combinación, aplicar el borde
                if not any(cell_ref in merge_range for merge_range in celdas_combinadas):
                    worksheet.write_blank(row, col, None, border_format)

        # Obtener el índice de la columna FECHA_NACIMIENTO
        if "FECHA_NACIMIENTO" in df_encabezado.columns:
            idx_fecha_nacimiento = df_encabezado.columns.get_loc("FECHA_NACIMIENTO")
            
            # Obtener las columnas que vienen después de FECHA_NACIMIENTO
            columnas_despues = df_encabezado.columns[idx_fecha_nacimiento + 1:]

            # Definir el formato (Arial 14, negrilla, centrado y alineado al medio)
            formato_texto = workbook.add_format({
                'bold': True,
                'font_name': 'Arial',
                'font_size': 14,
                'align': 'center',
                'valign': 'vcenter',
                'border': 1  # Borde en todas las direcciones
            })

            # Escribir los valores en worksheet a partir de K15
            col_inicio = 10  # K es la columna 10 (A=0, B=1, ..., K=10)
            fila_inicio = 14  # Fila 15 en Excel (0-index en Python)
            
            for i, columna in enumerate(columnas_despues):
                worksheet.write(fila_inicio, col_inicio + i, columna, formato_texto)  # Aplicar formato

        # ===================================================================
        # ===================================================================
        
        # Reemplazar NaN por cadena vacía para evitar el error
        df_plantilla = df_plantilla.fillna('')
        # Separar el valor de 'TIPODOC' por ':' y tomar solo el primer valor
        df_plantilla["TIPODOC"] = df_plantilla["TIPODOC"].str.split(":").str[0]

        # Crear un formato con Arial 14 y bordes
        formato_arial_borde = workbook.add_format({
            'font_name': 'Arial',
            'font_size': 14,
            'border': 1,  # Agregar borde en todas las direcciones
            'align': 'center',  # Alinear al centro
            'valign': 'vcenter' # Alinear verticalmente al centro
        })

        # Insertar valores de la columna 'NUMERO_REGISTRO' del DataFrame en la columna A, de la fila 17 a la 36
        for idx, doc_value in enumerate(df_plantilla['NUMERO_REGISTRO']):  
            row = 17 + idx  # Iniciar en la fila 17
            worksheet.write(row - 1, 0, doc_value, formato_arial_borde)  # Columna 'A' corresponde al índice 0

        # Insertar valores de la columna 'TIPODOC' del DataFrame en la columna B, de la fila 17 a la 36
        for idx, doc_value in enumerate(df_plantilla['TIPODOC']):  
            row = 17 + idx  # Iniciar en la fila 17
            worksheet.write(row - 1, 1, doc_value, formato_arial_borde)  # Columna 'B' corresponde al índice 1

        # Insertar valores de la columna 'DOC' del DataFrame en la columna C, de la fila 17 a la 36
        for idx, doc_value in enumerate(df_plantilla['DOC']):  
            row = 17 + idx  # Iniciar en la fila 17
            worksheet.write(row - 1, 2, doc_value, formato_arial_borde)  # Columna 'C' corresponde al índice 2
        
        # Insertar valores de la columna 'NOMBRE1' del DataFrame en la columna D, de la fila 17 a la 36
        for idx, doc_value in enumerate(df_plantilla['NOMBRE1']):  
            row = 17 + idx  # Iniciar en la fila 17
            worksheet.write(row - 1, 3, doc_value, formato_arial_borde)  # Columna 'D' corresponde al índice 3
        
        # Insertar valores de la columna 'NOMBRE2' del DataFrame en la columna E, de la fila 17 a la 36
        for idx, doc_value in enumerate(df_plantilla['NOMBRE2']):  
            row = 17 + idx  # Iniciar en la fila 17
            worksheet.write(row - 1, 4, doc_value, formato_arial_borde)  # Columna 'E' corresponde al índice 4
        
        # Insertar valores de la columna 'APELLIDO1' del DataFrame en la columna F, de la fila 17 a la 36
        for idx, doc_value in enumerate(df_plantilla['APELLIDO1']):  
            row = 17 + idx  # Iniciar en la fila 17
            worksheet.write(row - 1, 5, doc_value, formato_arial_borde)  # Columna 'F' corresponde al índice 5
        
        # Insertar valores de la columna 'APELLIDO2' del DataFrame en la columna G, de la fila 17 a la 36
        for idx, doc_value in enumerate(df_plantilla['APELLIDO2']):  
            row = 17 + idx  # Iniciar en la fila 17
            worksheet.write(row - 1, 6, doc_value, formato_arial_borde)  # Columna 'G' corresponde al índice 6
        
        # Insertar valores de la columna 'GENERO' del DataFrame en la columna H, de la fila 17 a la 36
        for idx, doc_value in enumerate(df_plantilla['GENERO']):  
            row = 17 + idx  # Iniciar en la fila 17
            worksheet.write(row - 1, 7, doc_value, formato_arial_borde)  # Columna 'H' corresponde al índice 7
        
        # Insertar valores de la columna 'GRADO_COD' del DataFrame en la columna I, de la fila 17 a la 36
        for idx, doc_value in enumerate(df_plantilla['GRADO_COD']):  
            row = 17 + idx  # Iniciar en la fila 17
            worksheet.write(row - 1, 8, doc_value, formato_arial_borde)  # Columna 'I' corresponde al índice 8
        
        # Insertar valores de la columna 'TIPO DE RACIÓN' del DataFrame en la columna J, de la fila 17 a la 36
        for idx, doc_value in enumerate(df_plantilla['TIPO DE RACIÓN']):  
            row = 17 + idx  # Iniciar en la fila 17
            worksheet.write(row - 1, 9, doc_value, formato_arial_borde)  # Columna 'J' corresponde al índice 9
        
        # ===========================================================
        # Codigo para el formato inferior de la planilla
        # ===========================================================
        # Combinar celdas A38 y C38 y agregar el texto en negrita
        worksheet.merge_range('A38:C38', 'RACIONES MENSUALES  PROGRAMADAS RPS', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_name': 'Arial',
                                'font_size': 11
                            }))

        # Combinar celdas E38 y F38 y agregar el texto en negrita
        worksheet.merge_range('E38:F38', 'RACIONES MENSUALES  ENTREGADAS RPS', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'right',
                                'valign': 'vcenter',
                                'font_name': 'Arial',
                                'font_size': 11
                            }))

        # Combinar celdas A40 y C40 y agregar el texto en negrita
        worksheet.merge_range('A40:C40', 'RACIONES MENSUALES PROGRAMADAS CCT', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_name': 'Arial',
                                'font_size': 11
                            }))

        # Combinar celdas E40 y F40 y agregar el texto en negrita
        worksheet.merge_range('E40:F40', 'RACIONES MENSUALES ENTREGADAS CCT', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'right',
                                'valign': 'vcenter',
                                'font_name': 'Arial',
                                'font_size': 11
                            }))

        # Combinar celdas A42 y C42 y agregar el texto en negrita
        worksheet.merge_range('A42:C42', 'RACIONES MENSUALES PROGRAMADAS RI', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_name': 'Arial',
                                'font_size': 11
                            }))

        # Combinar celdas E42 y F42 y agregar el texto en negrita
        worksheet.merge_range('E42:F42', 'RACIONES MENSUALES  ENTREGADAS RI', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'right',
                                'valign': 'vcenter',
                                'font_name': 'Arial',
                                'font_size': 11
                            }))

        # Combinar celdas A44 y C44 y agregar el texto en negrita
        worksheet.merge_range('A44:C44', 'NOMBRE DEL RESPONSABLE DEL OPERADOR', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_name': 'Arial',
                                'font_size': 12
                            }))

        # Combinar celdas G44 y P44 y agregar el texto en negrita
        worksheet.merge_range('G44:P44', 'NOMBRE RECTOR ESTABLECIMIENTO EDUCATIVO', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_name': 'Arial',
                                'font_size': 12
                            }))

        # Combinar celdas A46 y C46 y agregar el texto en negrita
        worksheet.merge_range('A46:C46', 'FIRMA DEL RESPONSABLE DEL OPERADOR', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_name': 'Arial',
                                'font_size': 12
                            }))

        # Combinar celdas G46 y P46 y agregar el texto en negrita
        worksheet.merge_range('G46:P46', 'FIRMA DEL RESPONSABLE DEL OPERADOR', 
                            workbook.add_format({
                                'bold': True,
                                'align': 'center',
                                'valign': 'vcenter',
                                'font_name': 'Arial',
                                'font_size': 12
                            }))

        # ===================================================================
        # Formato para borde inferior simple
        border_bottom_format = workbook.add_format({'bottom': 1})

        # Lista de celdas específicas a las que se aplicará el borde inferior
        celdas_borde_inferior = ["D38", "D40", "D42", "G38", "G40", "G42"]

        # Aplicar el formato a cada celda
        for celda in celdas_borde_inferior:
            worksheet.write(celda, None, border_bottom_format)
        # ===================================================================

        # Formato para borde inferior simple
        border_bottom_format = workbook.add_format({'bottom': 1})

        # Lista de rangos a los que se aplicará el borde inferior (excepto fila 41)
        rangos_borde_inferior = [
            (44, 3, 5),   # D44:F44 (fila 44, columnas 3 a 5)
            (44, 16, 33), # Q44:AH44 (fila 44, columnas 16 a 33)
            (46, 3, 5),   # D46:F46 (fila 46, columnas 3 a 5)
            (46, 16, 33)  # Q46:AH46 (fila 46, columnas 16 a 33)
        ]

        # Aplicar formato de borde inferior a las filas 47 y 49
        for fila, col_inicio, col_fin in rangos_borde_inferior:
            for col in range(col_inicio, col_fin + 1):
                worksheet.write_blank(fila - 1, col, None, border_bottom_format)

        # Formato para borde inferior grueso (solo A41:AI41)
        formato_borde_inferior = workbook.add_format({'bottom': 1})

        # Aplicar borde inferior grueso solo de A41 a AI41
        for col in range(0, 35):  # A (0) hasta AI (33)
            worksheet.write_blank(46, col, None, formato_borde_inferior)  # Fila 41 (índice 40 en Python)

        # Formato para la celda combinada en la fila 49
        formato_fila_49 = workbook.add_format({
            'bold': True,
            'align': 'left',      # Alineación a la izquierda
            'valign': 'top',      # Alineación en la parte superior
            'text_wrap': True,    # Ajuste de texto en la celda
            'font_name': 'Arial',
            'font_size': 12,
            'border': 1           # Bordes en todas las direcciones
        })

        # Combinar celdas A49:AI49 y aplicar formato
        worksheet.merge_range('A49:AI49', 'Observaciones:', formato_fila_49)

        # Formato para borde inferior grueso
        formato_borde_inferior_44 = workbook.add_format({'bottom': 1})

        # Aplicar borde solo en A44:AI44, evitando la intersección con la celda combinada en A49:AI49
        for col in range(0, 35):  # Columnas A (0) hasta AI (34)
            worksheet.write_blank(49, col, None, formato_borde_inferior_44)  # Fila 44 (índice 49 en Python)

        # Formato para borde izquierdo grueso
        formato_borde_izquierdo_AJ = workbook.add_format({'left': 2})  

        # Aplicar borde en la columna AJ desde la fila 1 hasta la 52
        for fila in range(0, 51):  # Filas 1 (0 en Python) hasta 52 (43 en Python)
            worksheet.write_blank(fila, 35, None, formato_borde_izquierdo_AJ)  # AJ es la columna 35 (A=0, B=1, ..., AJ=35)

        # Formato para borde inferior grueso
        formato_borde_inferior_44 = workbook.add_format({'bottom': 2})

        # Aplicar borde solo en A44:AI44, evitando la intersección con la celda combinada en A51:AI51
        for col in range(0, 35):  # Columnas A (0) hasta AI (34)
            worksheet.write_blank(50, col, None, formato_borde_inferior_44)  # Fila 44 (índice 51 en Python)
        # Guardar el archivo con las modificaciones
        writer.close()

    def separar_dataframes(self):
        # Leer el archivo Excel desde la celda A8 hasta AF27, tomando la fila 8 como encabezado
        # df_focalizacion = pd.read_excel(ruta_archivo_focalizacion)
        df_focalizacion = pd.read_excel(ruta_archivo_aplicacion_novedades)

        # Crear un diccionario para almacenar los DataFrames separados
        dfs_separados = {}

        # Agrupar por 'SEDE', 'JORNADA' y 'GRADO_COD'
        for (sede, jornada, grado), df_grupo in df_focalizacion.groupby(['SEDE', 'JORNADA', 'GRADO_COD']):
            # Agregar una columna de numeración secuencial para cada grupo
            df_grupo = df_grupo.copy()  # Evitar advertencias de SettingWithCopyWarning
            df_grupo['NUMERO_REGISTRO'] = range(1, len(df_grupo) + 1)

            # Dividir el grupo en bloques de 20 registros
            num_partes = (len(df_grupo) // 20) + (1 if len(df_grupo) % 20 > 0 else 0)

            for i in range(num_partes):
                # Extraer hasta 20 registros por fragmento
                df_parte = df_grupo.iloc[i * 20:(i + 1) * 20].copy()

                # Crear un nombre concatenado con la parte correspondiente
                nombre_df = f"{sede}_{jornada}_{grado}_parte_{i+1}"  # Se empieza en 1 en vez de 0
                # nombre_df = f"{sede}_{jornada}_{grado}"  # Se empieza en 1 en vez de 0
                
                # Guardarlo en el diccionario
                dfs_separados[nombre_df] = df_parte

        cantidad_dfs = len(dfs_separados)
        print(f"Cantidad de DataFrames generados: {cantidad_dfs}\n")

        df_encabezado = pd.read_excel(ruta_archivo_aplicacion_novedades, nrows=0)

        # Iterar sobre los DataFrames generados y pasarlos a la función
        for i, (nombre_df, df) in enumerate(dfs_separados.items(), start=1):
            print(f"Iteración {i}: Generando plantilla de control para: {nombre_df}")

            # Obtener el código DANE de la sede (suponiendo que está en el DataFrame)
            var_sede = df['SEDE'].iloc[0]  # Tomar el primer valor del grupo
            var_codigo_dane_sede = df['CODIGO_DANE_SEDE'].iloc[0]  # Tomar el primer valor del grupo
            var_jornada = df['JORNADA'].iloc[0]  # Tomar el primer valor del grupo
            var_grado = df['GRADO_COD'].iloc[0]  # Tomar el primer valor del grupo

            self.crear_plantilla_control(df, nombre_df, var_sede, var_codigo_dane_sede, var_jornada, df_encabezado, var_grado)

    def limpiar_carpeta_resultado(self, carpeta_resultado):

        # Eliminar y recrear la carpeta
        if os.path.exists(carpeta_resultado):
            shutil.rmtree(carpeta_resultado)  # Borra todo el contenido de una vez
        os.makedirs(carpeta_resultado)  # Crea la carpeta vacía nuevamente

        print(f"\nCarpeta {carpeta_resultado} vaciada y recreada.\n")

    def aplicar_novedades(self):
        # Leer insumo novedades
        df_novedades = pd.read_excel(ruta_archivo_novedades, sheet_name="Novedades")

        # Leer insumo focalizacion
        df_asistencia = pd.read_excel(ruta_archivo_focalizacion)

        # Obtener el índice de la columna FECHA_NACIMIENTO
        idx = df_asistencia.columns.get_loc("FECHA_NACIMIENTO")

        # Seleccionar las columnas que vienen después de FECHA_NACIMIENTO
        cols_despues = df_asistencia.columns[idx + 1:]

        # Rellenar todas las columnas siguientes con "X"
        df_asistencia[cols_despues] = "X"

        # Iterar sobre las novedades y aplicar los cambios
        for index, row in df_novedades.iterrows():
            fecha = row["FECHA"]
            sede = row["SEDE"]
            jornada = row["JORNADA"]
            grado = row["GRADO_COD"]
            novedad = row["TIPO_NOVEDAD"]
            detalle = row["DETALLE"]
        
            if pd.notna(detalle):  # Verifica que no sea NaN
                if isinstance(detalle, (int, float)) and not isinstance(detalle, bool):  
                    detalle = int(detalle)  # Convierte a entero si es numérico
                else:
                    detalle = str(detalle)  # Mantiene como texto si es un string
            else:
                detalle = None  # Si está vacío, asigna None

            # Filtrar la plantilla para encontrar los estudiantes correspondientes
            df_filtrado = df_asistencia.loc[
                (df_asistencia["SEDE"] == sede) & 
                (df_asistencia["JORNADA"] == jornada) & 
                (df_asistencia["GRADO_COD"] == grado)
            ]

            # Aplicar cambios según la novedad
            if novedad == "No hubo clases":
                # Obtener el número del día de la fecha
                num_dia = fecha.day  

                # Convertir col_dia a int para compararlo con las columnas del DataFrame
                col_dia = int(num_dia)

                # Verificar si la columna del día existe en el DataFrame
                if col_dia in [int(col) for col in df_asistencia.columns if str(col).isdigit()]:
                    # Dejar vacíos los valores en la columna correspondiente para los registros filtrados
                    df_asistencia.loc[df_filtrado.index, col_dia] = ""
                else:
                    print(f"Error: La columna '{col_dia}' no existe en el DataFrame.")
                    return
            
            elif novedad == "Asistencia parcial":
                # Numero de estudiantes
                num_estudiantes = df_filtrado.shape[0]

                # Validar que 'detalle' no supere el número de estudiantes
                if detalle > num_estudiantes:
                    print(f"""Error: La asistencia parcial supera la cantidad de estudiantes del grado.\n
                            (Numero de estudiantes: {num_estudiantes})\n en {sede}, {jornada}, {grado}.
                            (Asistencia parcial: {detalle})\n en {sede}, {jornada}, {grado}.""")
                    return  # Detener la ejecución
                else:
                    # Obtener el número del día de la fecha
                    num_dia = fecha.day  

                    # Convertir col_dia a int para compararlo con las columnas del DataFrame
                    col_dia = int(num_dia)

                    # Verificar si la columna del día existe en el DataFrame (conversión a int para evitar problemas con float64)
                    if col_dia not in [int(col) for col in df_asistencia.columns if str(col).isdigit()]:
                        print(f"Error: La columna '{col_dia}' no existe en el DataFrame.")
                        return

                    # Calcular cuántos valores deben quedar vacíos
                    num_vacios = num_estudiantes - detalle

                    # Seleccionar aleatoriamente índices de df_filtrado para dejarlos vacíos
                    indices_vacios = np.random.choice(df_filtrado.index, size=num_vacios, replace=False)

                    # Dejar vacíos los valores en los índices seleccionados
                    df_asistencia.loc[indices_vacios, col_dia] = ""

            elif novedad == "Cambio de complemento":
                pass  # Aquí se manejarían los totales al final, sin afectar las "X" individuales

        # Exportar el DataFrame modificado a un nuevo archivo Excel
        df_asistencia.to_excel(ruta_archivo_aplicacion_novedades, index=False)

        print("Archivo de Novedades aplicadas correctamente.")

    def concatenar_exceles(self):

        # Limpiar la carpeta 'Resultado' antes de generar los archivos
        self.limpiar_carpeta_resultado("Resultado excel")

        # Carpeta donde están los archivos
        carpeta = "Resultado"
        carpeta_salida = "Resultado excel"

        # Obtener la lista de archivos en la carpeta
        archivos_en_carpeta = [f for f in os.listdir(carpeta) if f.endswith(".xlsx")]

        # Diccionario para agrupar archivos por su nombre base (sin "_parte_X")
        grupos_archivos = defaultdict(list)
        
        for archivo in archivos_en_carpeta:
            # Tomar solo la parte antes del primer guion bajo
            nombre_base = archivo.split("_")[0]
            
            grupos_archivos[nombre_base].append(archivo)

        # Convertir a lista de listas
        listas_archivos = list(grupos_archivos.values())

        for grupo in listas_archivos:
            if grupo:  # Verificar que la lista no esté vacía

                # Ordenar la lista usando los números extraídos
                grupo = sorted(grupo, key=self.extraer_numeros)

                # Extraer el nombre base eliminando "_parte_X.xlsx"
                nombre_exportable = grupo[0].split("_")[0].strip()

                # Crear un nuevo libro de Excel
                nuevo_libro = Workbook()
                nuevo_libro.remove(nuevo_libro.active)  # Eliminar hoja por defecto

                # Copiar cada hoja con sus imágenes y estilos
                for i, archivo in enumerate(grupo, start=1):
                    ruta_archivo = os.path.join(carpeta, archivo)  # Ruta completa del archivo
                    wb_original = load_workbook(ruta_archivo)
                    hoja_original = wb_original.active  # Tomar la primera hoja

                    # Crear nueva hoja con el nombre correcto
                    nueva_hoja = nuevo_libro.create_sheet(title=f"Parte{i}")
                    nueva_hoja.sheet_view.showGridLines = False # Ocultar cuadrícula

                    # ✅ Copiar valores y estilos
                    for fila in hoja_original.iter_rows():
                        for celda in fila:
                            nueva_celda = nueva_hoja[celda.coordinate]
                            nueva_celda.value = celda.value  # Copiar valores

                            # ✅ Copiar estilos asegurando que no sean StyleProxy
                            if celda.has_style:
                                nueva_celda.font = copy(celda.font)
                                nueva_celda.fill = copy(celda.fill)
                                nueva_celda.border = copy(celda.border)
                                nueva_celda.alignment = copy(celda.alignment)
                                nueva_celda.number_format = celda.number_format
                                nueva_celda.protection = copy(celda.protection)

                    # ✅ Copiar dimensiones de columnas y filas
                    for col in hoja_original.column_dimensions:
                        nueva_hoja.column_dimensions[col] = copy(hoja_original.column_dimensions[col])

                    for row in hoja_original.row_dimensions:
                        nueva_hoja.row_dimensions[row] = copy(hoja_original.row_dimensions[row])

                    # ✅ Copiar celdas combinadas
                    for merge_range in hoja_original.merged_cells.ranges:
                        nueva_hoja.merge_cells(range_string=str(merge_range))

                    # ✅ Copiar imágenes
                    for img in hoja_original._images:
                        nueva_imagen = Image(img.ref)
                        nueva_imagen.anchor = img.anchor
                        nueva_hoja.add_image(nueva_imagen)

                    wb_original.close()

                # Guardar el nuevo archivo
                archivo_salida = os.path.join(carpeta_salida, f"{nombre_exportable}.xlsx")
                nuevo_libro.save(archivo_salida)

                print(f"✅ Archivo combinado guardado {nombre_exportable}")

    # Función para extraer los números de ordenamiento
    def extraer_numeros(self, nombre):
        match = re.search(r'_(\d+)_parte_(\d+)', nombre)
        if match:
            num = int(match.group(1))  # Número principal
            parte = int(match.group(2))  # Número de parte
            return (num, parte)
        return (float('inf'), float('inf'))  # Para evitar errores

    def main(self):
        # Iniciar proceso generacion de planillas
        print("="*60 + "\n  SE INICIA PROCESO DE GENERACION DE PLANILLAS DE CONTROL\n" + "="*60)

        # Registrar el tiempo de inicio
        inicio = time.time()

        # Aplicar las novedades al archivo de focalización
        self.aplicar_novedades()

        # Limpiar la carpeta 'Resultado' antes de generar los archivos
        self.limpiar_carpeta_resultado("Resultado")

        # Ejecutar la función
        self.separar_dataframes()

        time.sleep(2)

        # Concatenar los archivos generados
        self.concatenar_exceles()

        time.sleep(2)  # Esperar 2 segundos

        # Registrar el tiempo de fin
        fin = time.time()

        # Calcular el tiempo total de ejecución
        duracion = fin - inicio  # Diferencia en segundos
        horas, resto = divmod(duracion, 3600)  # Convertir a horas
        minutos, segundos = divmod(resto, 60)  # Convertir a minutos y segundos

        print(f"\nTiempo total de ejecución: {int(horas):02}:{int(minutos):02}:{int(segundos):02}")
        print("Proceso terminado\n")

if __name__ == "__main__":
    # Llamar a la función para ejecutar el proceso
    GeneradorPlantillas().main()
